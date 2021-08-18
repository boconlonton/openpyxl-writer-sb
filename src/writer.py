from openpyxl import Workbook


class ExcelFile:

    def __init__(self, name: str = 'user'):
        self.wb = Workbook()
        self.ws1 = self.wb.create_sheet('DATA')
        self.ws2 = self.wb.create_sheet('PROVINCE')
        self.ws2.sheet_properties.tabColor = "1072BA"
        self.ws3 = self.wb.create_sheet('DISTRICT')
        self.ws3.sheet_properties.tabColor = "1072BA"
        self.ws4 = self.wb.create_sheet('WARD')
        self.ws4.sheet_properties.tabColor = "1072BA"
        self.ws5 = self.wb.create_sheet('WORKING_STATUS')
        self.ws5.sheet_properties.tabColor = "1072BA"
        self.ws6 = self.wb.create_sheet('CONTRACT_TYPE')
        self.ws6.sheet_properties.tabColor = "1072BA"
        self.ws7 = self.wb.create_sheet('ETHNICS')
        self.ws7.sheet_properties.tabColor = "1072BA"
        self.ws8 = self.wb.create_sheet('RELIGION')
        self.ws8.sheet_properties.tabColor = "1072BA"
        self.ws9 = self.wb.create_sheet('TEACHING_TITLE')
        self.ws9.sheet_properties.tabColor = "1072BA"
        self.ws10 = self.wb.create_sheet('ACADEMIC_TITLE')
        self.ws10.sheet_properties.tabColor = "1072BA"
        self.ws11 = self.wb.create_sheet('DEGREE')
        self.ws11.sheet_properties.tabColor = "1072BA"
        self.file_name = name

    def write_province(self, data):
        province_headers = ['Mã Tỉnh/TP', 'Tên Tỉnh/TP']
        self.ws2.append(province_headers)
